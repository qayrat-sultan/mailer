import os
import glob
import re
from docx import Document
import openpyxl
import PyPDF2

# Regular expression pattern to match email addresses
email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'


# Function to extract emails from a Word document
def extract_emails_from_docx(file_path):
    doc = Document(file_path)
    emails = []
    for paragraph in doc.paragraphs:
        text = paragraph.text
        matches = re.findall(email_pattern, text)
        emails.extend(matches)
    return emails


# Function to extract emails from an Excel file
def extract_emails_from_xlsx(file_path):
    emails = []
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                text = str(cell.value)
                matches = re.findall(email_pattern, text)
                emails.extend(matches)
    return emails


# Function to extract emails from a PDF file
def extract_emails_from_pdf(file_path):
    emails = []
    with open(file_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        for page in reader.pages:
            text = page.extract_text()
            matches = re.findall(email_pattern, text)
            emails.extend(matches)
    return emails


# Function to find emails in a given file
def find_emails_in_file(file_path):
    extension = os.path.splitext(file_path)[1].lower()

    if extension == '.docx':
        return extract_emails_from_docx(file_path)
    elif extension == '.xlsx':
        return extract_emails_from_xlsx(file_path)
    elif extension == '.pdf':
        return extract_emails_from_pdf(file_path)
    elif extension == '.doc':
        # Handle .doc files if needed
        pass

    return []


# Function to find emails in a folder recursively
def find_emails_in_folder(folder_path):
    emails = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            emails.extend(find_emails_in_file(file_path))
    return emails


# Specify the folder path
folder_path = 'folder'

# Find all emails in the folder
all_emails = find_emails_in_folder(folder_path)

found_emails = []
# Print or use the extracted emails as desired
for email in all_emails:
    found_emails.append(email)
print(found_emails, )